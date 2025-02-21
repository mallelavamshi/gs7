import streamlit as st
import requests
import json
import os
import re
import anthropic
import pandas as pd
from PIL import Image
from io import BytesIO
from reportlab.platypus import Frame
from datetime import datetime, timedelta
import base64
from dotenv import load_dotenv
from auth_original import authenticated_layout
from database import (
    init_db, get_user_limits, increment_image_count, delete_user, 
    get_all_users, update_user_limit, is_admin, save_report, get_user_reports
)
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as PDFImage, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import mm
from streamlit.components.v1 import html
import time
import openpyxl
import random
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="EstateGenius AI", page_icon="🔍", layout="wide")

load_dotenv()
init_db()

ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY')
SEARCH_API_KEY = os.getenv('SEARCH_API_KEY')

def get_funny_message():
    """Return a random funny message for processing state"""
    messages = [
        "🎨 Teaching AI to appreciate art...",
        "📚 Consulting with virtual antique experts...",
        "💎 Calculating the value of priceless treasures...",
        "🔍 Decoding the mysteries of your items...",
        "✨ Making your items feel special...",
        "🤖 Asking AI what it thinks about your collection...",
        "🧐 Studying the fine details...",
        "📊 Comparing with millions of items...",
        "👨‍🎨 Getting opinions from digital experts...",
        "🔎 Making sure we don't miss anything...",
        "🎭 Analyzing artistic significance...",
        "📸 Enhancing image details...",
        "🏺 Consulting historical databases...",
        "🖼️ Examining craftsmanship...",
        "⚖️ Weighing market values..."
    ]
    return random.choice(messages)

def format_time(seconds):
    """Format seconds into HH:MM:SS"""
    if isinstance(seconds, (int, float)):
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        seconds = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return "00:00:00"

def create_pdf_report(results, output_file):
    """Create PDF report with images and analyses - modified for two columns"""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as PDFImage, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    
    class CustomDocTemplate(SimpleDocTemplate):
        def __init__(self, filename, **kwargs):
            super().__init__(filename, **kwargs)
            self.topMargin = 15*mm
            self.leftMargin = 25*mm
    
    doc = CustomDocTemplate(output_file, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Create custom style for analysis text with increased line spacing
    analysis_style = styles['BodyText']
    analysis_style.leading = 20  # Increase line spacing (default is usually around 12-14)
    
    # Create custom styles
    header_style = styles['Title']
    header_style.spaceAfter = 5
    
    contact_style = styles['Normal']
    contact_style.fontSize = 9
    contact_style.leading = 11
    contact_style.textColor = colors.gray
    
    tagline_style = styles['Normal']
    tagline_style.alignment = 1  # Center alignment
    tagline_style.fontSize = 11
    tagline_style.leading = 14
    tagline_style.spaceAfter = 20
    
    # Create header table with title and contact info
    header_style.textColor = colors.HexColor('#D97757')  # Set title color
    header_style.alignment = 1  # Center alignment
    
    # Contact info in the right column
    contact_info = [
        [Paragraph("Email: maggie@estategeniusai.com", contact_style)],
        [Paragraph("Mobile: (+1)469-659-7089", contact_style)],
        [Paragraph("Website: www.estategeniusai.com", contact_style)]
    ]
    
    # Title and taglines in the center column
    title_content = [
        [Paragraph("EstateGenius AI", header_style)],
        [Paragraph("Your Pricing Partner", tagline_style)],
        [Paragraph("Saves Hours of Internet Search", tagline_style)],
        [Paragraph("We Customize AI According to Your Needs", tagline_style)]
    ]
    
    # Create tables for each section
    contact_table = Table(contact_info, colWidths=[200])
    title_table = Table(title_content, colWidths=[300])
    
    # Style the tables
    contact_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    title_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    # Create a table for the header layout
    header_layout = Table([
        ['', title_table, contact_table]
    ], colWidths=[20, 300, 200])  # Added small left margin
    
    header_layout.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),  # Center title
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),   # Right align contact
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    elements.append(header_layout)
    elements.append(Spacer(1, 20))

    # Modified table with only two columns
    data = [["Image", "Analysis"]]  # Changed headers
    for result in results:
        try:
            img = PDFImage(result['temp_image_path'], width=150, height=150)
            # Use only the analysis
            row = [
                img,
                Paragraph(result['analysis'], analysis_style)
            ]
            data.append(row)
        except Exception as e:
            st.error(f"PDF error: {str(e)}")

    # Adjusted column widths for two columns
    table = Table(data, colWidths=[160, 420])  # Increased width for analysis column
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,0), 12),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,1), (1,-1), 'LEFT'),  # Left align analysis text
    ]))

    elements.append(table)

    try:
        doc.build(elements)
        return True
    except Exception as e:
        st.error(f"PDF creation failed: {str(e)}")
        return False

def create_excel_report(results, output_file):
    """Create Excel report with images and analyses - modified for two columns"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add contact information (top left)
    ws['A1'] = "Email: maggie@estategeniusai.com"
    ws['A2'] = "Mobile: (+)469-659-7089"
    ws['A3'] = "Website: www.estategeniusai.com"
    
    # Style contact info
    for cell in [ws['A1'], ws['A2'], ws['A3']]:
        cell.font = openpyxl.styles.Font(size=9, color="666666")  # Gray color
        cell.alignment = openpyxl.styles.Alignment(vertical='center')
    
    # Add header and taglines (center aligned in column B with same width as analysis)
    header_cell = ws['B1']
    header_cell.value = "EstateGenius AI"
    header_cell.font = openpyxl.styles.Font(size=16, bold=True)
    header_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    # Add taglines
    taglines = [
        "Your Pricing Partner",
        "Saves Hours of Internet Search",
        "We Customize AI According to Your Needs"
    ]
    
    for idx, tagline in enumerate(taglines, 2):
        cell = ws[f'B{idx}']
        cell.value = tagline
        cell.font = openpyxl.styles.Font(size=11)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    # Add some space before the table headers
    start_row = 6  # Start the actual content from row 6
    
    # Modified headers for two columns
    headers = ['Image', 'Analysis']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Add data with combined filename and analysis
    for row_idx, result in enumerate(results, start_row + 1):
        try:
            if os.path.exists(result['temp_image_path']):
                img = XLImage(result['temp_image_path'])
                img.width = 200
                img.height = 200
                ws.add_image(img, f'A{row_idx}')
            
            # Use only the analysis in second column
            analysis_cell = ws.cell(row=row_idx, column=2, value=result['analysis'])
            analysis_cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
            
            # Set row height based on content
            ws.row_dimensions[row_idx].height = max(150, len(result['analysis'].split('\n')) * 15)
            
        except Exception as e:
            st.error(f"Excel error: {str(e)}")

    # Adjust column widths for two columns
    ws.column_dimensions['A'].width = 30  # For images
    ws.column_dimensions['B'].width = 70  # Wider column for combined analysis

    # Set row heights for header section
    for i in range(1, 5):  # Rows 1-4 (contact info and taglines)
        ws.row_dimensions[i].height = 20
    
    # Add borders to content cells
    content_range = f'A{start_row}:B{len(results) + start_row}'
    for row in ws[content_range]:
        for cell in row:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )

    try:
        wb.save(output_file)
        return True
    except Exception as e:
        st.error(f"Save error: {str(e)}")
        return False
    
def create_basic_report(images):
    """Create report data without API processing and analysis"""
    results = []
    temp_files = []
    
    for image in images:
        try:
            response = requests.get(image['url'])
            if response.status_code != 200:
                continue
                
            img_path = f"temp_{image['id']}.jpg"
            with Image.open(BytesIO(response.content)) as img:
                img.convert('RGB').save(img_path)
            temp_files.append(img_path)

            results.append({
                'name': image['name'],
                'temp_image_path': img_path,
                'analysis': ''
            })
            
        except Exception as e:
            st.error(f"Basic processing error: {str(e)}")
            
    return results, temp_files

def get_anthropic_analysis(json_data):
    """Get analysis from Anthropic API"""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    prompt = f"""Analyze product search results and provide structured summary following these guidelines:
    1. Name: If there are multiple listings with same name or almost similar name then the item must be exactly 
    the same item as that in image. then assertively say the item: "Name", if the all the names in item listings  are mutually exclusive
    then the first listing is likely the item similar to the image then say item: "likely- first listing item name"
    2.opinion: tell succintly what you know about the item, its collector market and trends.
    3. ebay prices: give the prices seen in the all the ebay listings seperated by commas, just the prices
    4. etsy prices:give the prices seen in the all the etsy listings seperated by commas, just the prices
    5. amazon,walmart,macys prices if available.
    5. auctions houses:just say this item was or is listed in this action houses but dont say the prices in there.
    6. give all the above bullet points for clear reading.
    7. dont give any introduction like this:"Here's the structured summary:"
    

    Data: {json.dumps(json_data, indent=2)}"""

    try:
        message = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}]
        )
        return message.content[0].text if message.content else "No analysis generated"
    except Exception as e:
        st.error(f"Analysis error: {str(e)}")
        return "Analysis failed"

def search_google_lens(image_url):
    """Search Google Lens for image matches"""
    try:
        response = requests.get(
            "https://www.searchapi.io/api/v1/search",
            params={
                "engine": "google_lens",
                "url": image_url,
                "api_key": SEARCH_API_KEY
            }
        )
        return response.json().get("visual_matches", [])[:15]
    except Exception as e:
        st.error(f"Lens search failed: {str(e)}")
        return []

def extract_file_ids_from_folder(folder_url):
    """Extract file IDs from Google Drive folder"""
    try:
        folder_id = folder_url.split('/')[-1]
        files_url = f"https://drive.google.com/drive/folders/{folder_id}"
        response = requests.get(files_url)
        
        pattern = r"https://drive\.google\.com/file/d/([a-zA-Z0-9_-]+)"
        file_ids = list(set(re.findall(pattern, response.text)))
        
        return [{'id': fid, 'url': f"https://drive.google.com/uc?id={fid}", 'name': f"image_{fid}.jpg"} for fid in file_ids]
    except Exception as e:
        st.error(f"Error extracting files: {str(e)}")
        return []

def admin_panel():
    """Admin dashboard functionality"""
    st.header("🛠️ Admin Dashboard")
    st.subheader("User Management")
    users = get_all_users()
    df = pd.DataFrame(users, columns=["ID", "Username", "Email", "Role", "Max Images", "Processed Images"])
    st.dataframe(df)

    st.markdown("---")
    st.subheader("Delete User")
    del_id = st.number_input("Enter User ID to delete", min_value=1)
    if st.button("Delete User"):
        if delete_user(del_id):
            st.success("User deleted successfully")
            st.rerun()
        else:
            st.error("Failed to delete user")

    st.markdown("---")
    st.subheader("Set Processing Limits")
    limit_user_id = st.number_input("User ID for limit update", min_value=1)
    new_limit = st.number_input("New maximum images", min_value=1, value=100)
    if st.button("Update Limit"):
        if update_user_limit(limit_user_id, new_limit):
            st.success("Image limit updated")
            st.rerun()
        else:
            st.error("Failed to update limit")

def main_application():
    """Main application logic"""
    st.title("🔍 EstateGenius AI")
    st.markdown("---")

    with st.sidebar:
        st.header("User Controls")
        st.write(f"Logged in as: **{st.session_state.authenticated_user}**")
        
        if is_admin(st.session_state.authenticated_user):
            if st.checkbox("Show Admin Panel"):
                admin_panel()
                return
        
        current_count, max_allowed = get_user_limits(st.session_state.authenticated_user)
        st.metric("Processed Images", f"{current_count}/{max_allowed}")
        st.markdown("---")
        st.button("Logout", on_click=lambda: st.session_state.pop("authenticated_user"), key="logout_button")

        st.subheader("📚 Past Reports")
        reports = get_user_reports(st.session_state.authenticated_user)
        
        # Group reports by timestamp
        report_groups = {}
        for report in reports:
            try:
                full_path = report[0]
                if not full_path.startswith('/var/lib/estateai/reports'):
                    full_path = os.path.join('/var/lib/estateai/reports', os.path.basename(full_path))
                
                base_name = os.path.basename(full_path).split('.')[0]
                parts = base_name.split('_')
                if len(parts) >= 3:
                    timestamp_str = f"{parts[1]}_{parts[2]}"
                    report_groups.setdefault(timestamp_str, []).append(full_path)
            except Exception as e:
                st.error(f"Error processing report {full_path}: {str(e)}")

        for timestamp_str, report_paths in report_groups.items():
            try:
                report_date = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                with st.expander(f"📅 {report_date.strftime('%Y-%m-%d %H:%M')}"):
                    for path in report_paths:
                        try:
                            if path.endswith('.pdf'):
                                btn_label = "📄 PDF Version"
                                mime_type = "application/pdf"
                            elif path.endswith('.xlsx'):
                                btn_label = "📊 Excel Version"
                                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            else:
                                continue
                            
                            with open(path, "rb") as f:
                                st.download_button(
                                    label=btn_label,
                                    data=f.read(),
                                    file_name=os.path.basename(path),
                                    mime=mime_type
                                )
                        except Exception as e:
                            st.error(f"Error loading file {path}: {str(e)}")
            except Exception as e:
                st.error(f"Error loading report {timestamp_str}: {str(e)}")

    folder_url = st.text_input("Google Drive Folder URL", 
                              placeholder="https://drive.google.com/drive/folders/...")

    col1, col2 = st.columns(2)
    with col1:
        process_button = st.button("Process Images with Analysis", type="primary")
    with col2:
        basic_process_button = st.button("Process without Appraisal", type="secondary")

    if process_button or basic_process_button:
        if not folder_url:
            st.error("Please enter a valid folder URL")
            return

        with st.status("🔍 Processing images...", expanded=True) as status:
            # Initialize containers for updates
            progress_container = st.container()
            metrics_container = st.container()
            message_container = st.empty()
            
            with progress_container:
                st.write("Initializing...")
                progress_bar = st.progress(0)
                status_text = st.empty()
            
            with metrics_container:
                col1, col2 = st.columns(2)
                elapsed_placeholder = col1.empty()
                remaining_placeholder = col2.empty()
            
            # Get images from folder
            images = extract_file_ids_from_folder(folder_url)
            image_count = len(images)
            
            if image_count > 25:
                st.warning("Processing first 25 images of the folder")
                images = images[:25]

            current_count, max_allowed = get_user_limits(st.session_state.authenticated_user)
            if current_count + image_count > max_allowed:
                st.error(f"Image limit exceeded: {current_count + image_count}/{max_allowed}")
                return

            # Start time tracking
            start_time = time.time()
            
            if basic_process_button:
                results, temp_files = create_basic_report(images)
                status.update(label="📄 Creating basic reports...", state="running")
            else:
                results = []
                temp_files = []
                
                for idx, image in enumerate(images, 1):
                    try:
                        # Update progress
                        progress = idx / len(images)
                        progress_bar.progress(progress)
                        
                        # Update status text with spinner and image count
                        spinner_chars = ["⠋", "⠙", "⠹", "⠸", "⠼", "⠴", "⠦", "⠧", "⠇", "⠏"]
                        spinner = spinner_chars[int(time.time() * 10) % len(spinner_chars)]
                        status_text.write(f"{spinner} Processing image {idx} of {len(images)}")
                        
                        # Calculate times
                        elapsed_time = time.time() - start_time
                        elapsed_str = format_time(elapsed_time)
                        
                        if idx > 1:
                            avg_time_per_image = elapsed_time / (idx - 1)
                            remaining_images = len(images) - idx + 1
                            estimated_remaining = avg_time_per_image * remaining_images
                            remaining_str = format_time(estimated_remaining)
                        else:
                            remaining_str = "Calculating..."
                        
                        # Update metrics with icons
                        elapsed_placeholder.metric("⏱️ Elapsed Time", elapsed_str)
                        remaining_placeholder.metric("⏳ Estimated Remaining", remaining_str)
                        
                        # Show funny message with periodic updates
                        if idx % 2 == 0 or idx == 1:  # Update message every 2 images or on first image
                            message_container.info(get_funny_message())
                        
                        # Process image
                        response = requests.get(image['url'])
                        if response.status_code != 200:
                            continue
                            
                        img_path = f"temp_{image['id']}.jpg"
                        with Image.open(BytesIO(response.content)) as img:
                            img.convert('RGB').save(img_path)
                        temp_files.append(img_path)

                        lens_results = search_google_lens(image['url'])
                        if lens_results:
                            analysis = get_anthropic_analysis(lens_results)
                            results.append({
                                'name': image['name'],
                                'temp_image_path': img_path,
                                'analysis': analysis
                            })
                        
                        # Small delay to allow UI updates
                        time.sleep(0.1)

                    except Exception as e:
                        st.error(f"Image {idx} error: {str(e)}")

                # Clear message container when done
                message_container.empty()

            if results:
                base_name = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                # Update the reports path to use the persistent volume
                reports_dir = "/var/lib/estateai/reports"
                os.makedirs(reports_dir, exist_ok=True)
                
                pdf_report_name = os.path.join(reports_dir, f"{base_name}.pdf")
                excel_report_name = os.path.join(reports_dir, f"{base_name}.xlsx")
                
                # Show generating reports message
                status.update(label="📊 Generating reports...", state="running")
                
                pdf_success = create_pdf_report(results, pdf_report_name)
                excel_success = create_excel_report(results, excel_report_name)
                
                if pdf_success and excel_success:
                    save_report(st.session_state.authenticated_user, pdf_report_name)
                    save_report(st.session_state.authenticated_user, excel_report_name)
                    increment_image_count(st.session_state.authenticated_user, image_count)
                    
                    status.update(label="✅ Processing complete!", state="complete")
                    
                    # Success message with download buttons
                    st.success("Reports generated successfully!")
                    
                    with open(pdf_report_name, "rb") as f_pdf, open(excel_report_name, "rb") as f_xlsx:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.download_button(
                                label="📥 Download PDF Report",
                                data=f_pdf.read(),
                                file_name=os.path.basename(pdf_report_name),
                                mime="application/pdf",
                                key="pdf_download"
                            )
                        with col2:
                            st.download_button(
                                label="📥 Download Excel Report",
                                data=f_xlsx.read(),
                                file_name=os.path.basename(excel_report_name),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="excel_download"
                            )

            # Cleanup temp files
            for temp_file in temp_files:
                try:
                    os.remove(temp_file)
                except Exception as e:
                    st.warning(f"Failed to remove temporary file {temp_file}: {str(e)}")

if __name__ == "__main__":
    authenticated_layout(main_application)